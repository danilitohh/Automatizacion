"""Lectura de Excel exclusiva para Bot Leads Deploy.

Este servicio NO depende de la columna "Activo de Test". Las URLs de los
programas se obtienen del catálogo interno backend/data/utel_programas1.xlsx
durante la preparación del lote.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import load_workbook

from .bot_spreadsheet_service import BotSpreadsheetService


class LeadsDeploySpreadsheetService(BotSpreadsheetService):
    """Parser específico para la matriz Leads Deploy."""

    def preview(self, content: bytes, filename: str) -> dict[str, Any]:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
        sheets: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = self._header_index(values)
            if header_index is None:
                continue

            headers = [self._text(value) for value in values[header_index]]
            detected_mapping = self._mapping(headers)

            # Para Leads Deploy solo son imprescindibles estas columnas.
            if not all(
                key in detected_mapping
                for key in ("country", "level", "form_type")
            ):
                continue

            # "Activo de Test" se ignora a propósito. El catálogo interno decide
            # programa + URL en el worker.
            detected_mapping.pop("utel_url", None)

            normalized_rows: list[dict[str, Any]] = []
            last_country = ""

            for row_number, row_values in enumerate(
                values[header_index + 1 :],
                header_index + 2,
            ):
                if not any(self._text(value) for value in row_values):
                    continue

                item = {
                    key: self._text(row_values[index])
                    if index < len(row_values)
                    else ""
                    for key, index in detected_mapping.items()
                }

                current_country = item.get("country", "")
                if current_country:
                    last_country = current_country
                else:
                    current_country = last_country

                level = item.get("level", "")
                form_type = self._normalize_form_type(
                    item.get("form_type", "")
                )

                if not current_country or not level or not form_type:
                    continue

                item.update(
                    {
                        "country": current_country,
                        "level": level,
                        "form_type": form_type,
                        "utel_url": "",
                        "workflow_mode": "form_validation",
                    }
                )
                normalized_rows.append(
                    {"row_number": row_number, **item}
                )

            if not normalized_rows:
                continue

            sheets.append(
                {
                    "name": worksheet.title,
                    "headers": headers,
                    "mapping": {
                        key: headers[index]
                        for key, index in detected_mapping.items()
                    },
                    "rows": normalized_rows[:200],
                }
            )

        suggestions = [
            (
                f"{sheet['name']}: {len(sheet['rows'])} casos Leads Deploy "
                "detectados. Activo de Test se ignora; las URLs se toman "
                "del catálogo interno."
            )
            for sheet in sheets
        ]

        return {
            "filename": filename,
            "sheets": sheets,
            "suggestions": suggestions,
        }

    def rows_for_mapping(
        self,
        content: bytes,
        mapping: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Lee casos Deploy sin exigir URL de programa en el Excel."""

        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
        selected = {
            key: self._normalize(value)
            for key, value in mapping.items()
            if isinstance(value, str) and value
        }

        required_mapping = ("country", "level", "form_type")
        if not all(selected.get(key) for key in required_mapping):
            return []

        rows: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = self._header_index(values)
            if header_index is None:
                continue

            headers = [self._text(value) for value in values[header_index]]
            indexes = {
                key: next(
                    (
                        index
                        for index, header in enumerate(headers)
                        if self._normalize(header) == selected_value
                    ),
                    None,
                )
                for key, selected_value in selected.items()
            }

            if any(indexes.get(key) is None for key in required_mapping):
                continue

            last_country = ""

            for row_number, row_values in enumerate(
                values[header_index + 1 :],
                header_index + 2,
            ):
                level = self._cell(row_values, indexes.get("level"))
                form_type = self._normalize_form_type(
                    self._cell(row_values, indexes.get("form_type"))
                )
                current_country = self._cell(
                    row_values,
                    indexes.get("country"),
                )

                if current_country:
                    last_country = current_country
                else:
                    current_country = last_country

                if not level or not form_type or not current_country:
                    continue

                program = self._cell(
                    row_values,
                    indexes.get("program_name"),
                )

                rows.append(
                    {
                        "sheet": worksheet.title,
                        "row_number": row_number,
                        "program_name": program,
                        "level": level,
                        "modality": self._cell(
                            row_values,
                            indexes.get("modality"),
                        ),
                        "country": current_country,
                        "form_type": form_type,
                        "inconcert_url": self._cell(
                            row_values,
                            indexes.get("inconcert_url"),
                        ),
                        "lead_origin_url": self._cell(
                            row_values,
                            indexes.get("lead_origin_url"),
                        ),
                        # Intencionalmente vacío: el worker usa
                        # choose_catalog_program() antes de validar UtelQaConfig.
                        "utel_url": "",
                        "workflow_mode": "form_validation",
                        "test_case": level or program,
                    }
                )

        return rows
