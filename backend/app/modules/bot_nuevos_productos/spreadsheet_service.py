"""Importa filas de Excel para preparar ejecuciones del Bot sin columnas fijas."""

import io
import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .program_rotation_service import ProgramRotationService


class BotSpreadsheetService:
    DEFAULT_CATALOG_PATH = Path(__file__).resolve().parents[3] / "data" / "Programas_UTEL_Todos_los_Paises.xlsx"
    # Agrega nuevos paises aqui cuando se incorporen nuevos balanceadores.
    INCONCERT_BY_COUNTRY = {
        "mexico": "https://mas-utel.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "argentina": "https://mas-utel-arg.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "colombia": "https://mas-utel-col.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "peru": "https://mas-utel-pe.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "ecuador": "https://mas-utel-ec.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "bolivia": "https://mas-utel-bol.inconcertcc.com/mas/contact/people/view/185069",
        "united states": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "usa": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "estados unidos": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "chile": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "paraguay": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "guatemala": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "panama": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "el salvador": "https://mas-utel-emergentes.inconcertcc.com/mas/contact/people",
        "dominicana": "https://mas-utel-dom.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "republica dominicana": "https://mas-utel-dom.inconcertcc.com/login?redirect=%2Fmas%2Fhome",
        "filipinas": "https://mas-utel-singapur.infunnel.inconcert.cloud/",
    }

    def __init__(self, catalog_path: Path | str | None = None):
        self.catalog_path = Path(catalog_path) if catalog_path else self.DEFAULT_CATALOG_PATH

    def catalog_programs(self, country: str, level: str, modality: str = "") -> list[dict[str, str]]:
        """Devuelve programas oficiales que coinciden con país, nivel y modalidad."""

        if not self.catalog_path.is_file():
            return []
        country_key = self._catalog_key(country)
        level_key = self._catalog_level_key(level)
        modality_key = self._catalog_modality_key(modality)
        workbook = load_workbook(self.catalog_path, read_only=True, data_only=True)
        matches: list[dict[str, str]] = []
        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = next(
                (i for i, row in enumerate(values[:20]) if self._catalog_header_row(row)),
                None,
            )
            if header_index is None:
                continue
            headers = [self._normalize(self._text(v)) for v in values[header_index]]
            indexes = {
                "country": self._catalog_index(headers, ("pais", "country")),
                "modality": self._catalog_index(headers, ("modalidad", "modality")),
                "level": self._catalog_index(headers, ("nivel", "level")),
                "program": self._catalog_index(headers, ("programa", "program")),
                "url": self._catalog_index(headers, ("url del programa", "program url", "url")),
            }
            if indexes["program"] is None or indexes["url"] is None:
                continue
            for row in values[header_index + 1 :]:
                row_country = self._cell(row, indexes["country"])
                # Si no existe columna de país, el catálogo lo indica en la hoja.
                if indexes["country"] is None:
                    row_country = worksheet.title
                row_level = self._cell(row, indexes["level"])
                row_modality = self._cell(row, indexes["modality"])
                program = self._cell(row, indexes["program"])
                url = self._cell(row, indexes["url"])
                if not program or not url:
                    continue
                if self._catalog_key(row_country) != country_key:
                    continue
                if not self._catalog_level_matches(level_key, self._catalog_level_key(row_level)):
                    continue
                if modality_key and not self._catalog_modality_matches(modality_key, self._catalog_modality_key(row_modality)):
                    continue
                matches.append({"text": program, "url": url})
        # El libro puede contener registros repetidos entre secciones; la URL
        # y el nombre forman una identidad estable para la rotación.
        return list({(item["text"], item["url"]): item for item in matches}.values())

    def choose_catalog_program(self, country: str, level: str, modality: str, database_path: Path) -> dict[str, str] | None:
        """Selecciona el siguiente programa oficial sin depender del menú web."""

        candidates = self.catalog_programs(country, level, modality)
        if not candidates:
            return None
        return ProgramRotationService(database_path).choose(
            ["catalog", country, level, modality], candidates
        )

    @classmethod
    def _catalog_header_row(cls, row: tuple[Any, ...]) -> bool:
        normalized = {cls._normalize(cls._text(value)) for value in row}
        return bool({"programa", "program"} & normalized) and bool(
            {"url del programa", "program url", "url"} & normalized
        )

    @staticmethod
    def _catalog_index(headers: list[str], aliases: tuple[str, ...]) -> int | None:
        return next((i for i, header in enumerate(headers) if header in aliases), None)

    @classmethod
    def _catalog_key(cls, value: str) -> str:
        key = cls._normalize(value)
        if key in {"rep. dominicana", "republica dominicana", "dominicana"}:
            return "dominicana"
        return {"mexico": "mexico", "méxico": "mexico", "peru": "peru", "perú": "peru", "panama": "panama", "panamá": "panama", "philippines": "philippines", "filipinas": "philippines"}.get(key, key)

    @classmethod
    def _catalog_level_key(cls, value: str) -> str:
        key = cls._normalize(value)
        if "master" in key or "magister" in key or "maestr" in key:
            return "master"
        if "licenc" in key or "bachelor" in key or "carrera" in key:
            return "bachelor"
        if "doctor" in key:
            return "doctor"
        if "diplom" in key:
            return "diplomado"
        if "bootcamp" in key:
            return "bootcamp"
        if "bachiller" in key:
            return "bachillerato"
        return key

    @classmethod
    def _catalog_modality_key(cls, value: str) -> str:
        key = cls._normalize(value)
        if "ejecut" in key:
            return "ejecutiva"
        if "hibr" in key:
            return "hibrida"
        if "linea" in key or "online" in key or "virtual" in key:
            return "online"
        return key

    @staticmethod
    def _catalog_level_matches(expected: str, actual: str) -> bool:
        return expected == actual or not expected

    @staticmethod
    def _catalog_modality_matches(expected: str, actual: str) -> bool:
        return expected == actual or not actual

    @classmethod
    def effective_country(cls, country: str, level: str, url: str) -> str:
        # Global no identifica un CRM; resolver solo con evidencia explicita.
        if cls._normalize(country) == "global" and (
            re.search(r"\b(filipinas|philippines)\b", cls._normalize(level))
            or re.search(r"/philippines(?:/|\?|$)", url, re.I)
        ):
            return "Filipinas"
        return country

    @classmethod
    def default_inconcert_url(cls, country: str) -> str:
        normalized = cls._normalize(country)
        return cls.INCONCERT_BY_COUNTRY.get(normalized, "")

    """Detecta encabezados por significado y devuelve filas normalizadas."""

    ALIASES = {
        "country": ("country", "pais", "país", "locale"),
        "level": ("nivel", "level", "grado"),
        "modality": ("modalidad", "modality"),
        "utel_url": ("activo de test", "url utel", "utel url", "url nueva", "url page", "url"),
        "inconcert_url": (
            "url inconcert/balanceador",
            "inconcert/balanceador",
            "url inconcert",
            "inconcert url",
            "balanceador",
        ),
        "lead_origin_url": (
            "url origen lead",
            "url origen del lead",
            "origen lead",
        ),
        "form_type": ("location", "formulario", "form type", "ubicacion", "ubicación"),
        "program_name": ("programa", "program", "carrera", "producto"),
        "lead_name": ("nombre", "name", "responsable"),
        "lead_email": ("email", "correo"),
        "lead_phone": ("phone", "telefono", "teléfono", "celular"),
    }

    def preview(self, content: bytes, filename: str) -> dict[str, Any]:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            header_index = self._header_index(rows)
            if header_index is None:
                continue
            headers = [self._text(value) for value in rows[header_index]]
            mapping = self._mapping(headers)
            page_url_mode = any(self._normalize(header) == "url page" for header in headers)
            normalized_rows = []
            for row_number, values in enumerate(rows[header_index + 1 :], header_index + 2):
                if not any(self._text(value) for value in values):
                    continue
                item = {key: self._text(values[index]) if index < len(values) else "" for key, index in mapping.items()}
                if page_url_mode and "form_type" not in item:
                    item["form_type"] = "tarjeta"
                if item.get("utel_url") or item.get("program_name") or item.get("country"):
                    normalized_rows.append({"row_number": row_number, **item})
            sheets.append({"name": worksheet.title, "headers": headers, "mapping": {key: headers[index] for key, index in mapping.items()}, "rows": normalized_rows[:200]})
        return {"filename": filename, "sheets": sheets, "suggestions": self._suggestions(sheets)}

    def rows_for_mapping(self, content: bytes, mapping: dict[str, str]) -> list[dict[str, Any]]:
        """Lee todas las filas usando los encabezados elegidos por el usuario."""

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        selected = {key: self._normalize(value) for key, value in mapping.items() if value}
        workflow_mode = (
            "form_validation"
            if selected.get("level")
            and selected.get("form_type")
            and selected.get("country")
            and not selected.get("program_name")
            else "product_release"
        )
        rows = []
        for worksheet in workbook.worksheets:
            last_country = ""
            values = list(worksheet.iter_rows(values_only=True))
            header_index = self._header_index(values)
            if header_index is None:
                continue
            headers = [self._text(value) for value in values[header_index]]
            indexes = {
                key: next((index for index, header in enumerate(headers) if self._normalize(header) == selected_value), None)
                for key, selected_value in selected.items()
            }
            if indexes.get("utel_url") is None or (indexes.get("program_name") is None and indexes.get("level") is None):
                continue
            for row_number, row_values in enumerate(values[header_index + 1 :], header_index + 2):
                program = self._cell(row_values, indexes.get("program_name"))
                level = self._cell(row_values, indexes.get("level"))
                url = self._cell(row_values, indexes["utel_url"])
                if not (program or level) or not url:
                    continue
                current_country = self._cell(row_values, indexes.get("country"))
                if current_country:
                    last_country = current_country
                else:
                    current_country = last_country
                rows.append({
                    "sheet": worksheet.title,
                    "row_number": row_number,
                    "program_name": program,
                    "level": level,
                    "modality": self._cell(row_values, indexes.get("modality")),
                    "country": current_country,
                    "form_type": self._normalize_form_type(self._cell(row_values, indexes.get("form_type"))),
                    "inconcert_url": self._cell(row_values, indexes.get("inconcert_url")),
                    "lead_origin_url": self._cell(row_values, indexes.get("lead_origin_url")),
                    "utel_url": url,
                    "workflow_mode": workflow_mode,
                    "test_case": level or program,
                })
        return rows

    @classmethod
    def deploy_navigation_plan(cls, raw_level: str, country: str) -> dict[str, str]:
        """Convierte el nivel descriptivo de Leads Deploy en opciones reales del menu/formulario."""

        level = cls._normalize(raw_level)
        if "filipinas master" in level or "philippines master" in level:
            return {"modality": "Online", "level": "Master's Degree", "navigation_modality": "", "navigation_level": "Master", "navigation_sublevel": ""}
        if "filipinas bachelor" in level or "philippines bachelor" in level or "india bachelor" in level:
            return {"modality": "Online", "level": "Bachelor's Degree", "navigation_modality": "", "navigation_level": "Bachelor", "navigation_sublevel": ""}
        if "ejecutiva" in level:
            base = "Maestria" if "maestr" in level else "Licenciatura"
            return {"modality": "Ejecutiva", "level": base, "navigation_modality": "Modalidad ejecutiva", "navigation_level": f"{base}s", "navigation_sublevel": ""}
        if "hibrida" in level:
            base = "Maestria" if "maestr" in level else "Licenciatura"
            return {"modality": "Hibrida", "level": base, "navigation_modality": "Modalidad hibrida", "navigation_level": f"{base}s", "navigation_sublevel": ""}
        if "diplom" in level:
            return {"modality": "En linea", "level": "Diplomado", "navigation_modality": "Modalidad en linea", "navigation_level": "Educacion Continua", "navigation_sublevel": "Diplomados"}
        if "bootcamp" in level:
            return {"modality": "En linea", "level": "Bootcamp", "navigation_modality": "Modalidad en linea", "navigation_level": "Educacion Continua", "navigation_sublevel": "Bootcamps"}
        if "bachiller" in level:
            return {"modality": "En linea", "level": "Bachillerato", "navigation_modality": "Modalidad en linea", "navigation_level": "Bachillerato", "navigation_sublevel": ""}
        if "doble" in level and ("usa" in level or "mex" in level):
            return {"modality": "En linea", "level": "Licenciatura", "navigation_modality": "Modalidad en linea", "navigation_level": "Licenciaturas", "navigation_sublevel": "Doble titulacion Mex-USA"}
        if "ingl" in level:
            return {"modality": "En linea", "level": "Maestria", "navigation_modality": "Modalidad en linea", "navigation_level": "Masteres Internacionales", "navigation_sublevel": ""}
        if "doctor" in level:
            return {"modality": "En linea", "level": "Doctorado", "navigation_modality": "Modalidad en linea", "navigation_level": "Doctorados", "navigation_sublevel": ""}
        if "maestr" in level:
            return {"modality": "En linea", "level": "Maestria", "navigation_modality": "Modalidad en linea", "navigation_level": "Maestrias", "navigation_sublevel": ""}
        return {"modality": "En linea", "level": "Licenciatura", "navigation_modality": "Modalidad en linea", "navigation_level": "Licenciaturas", "navigation_sublevel": ""}

    @staticmethod
    def _cell(values: tuple[Any, ...], index: int | None) -> str:
        return str(values[index]).strip() if index is not None and index < len(values) and values[index] is not None else ""

    @staticmethod
    def _normalize_form_type(value: str) -> str:
        """Convierte nombres visibles del Excel a los valores internos del bot."""

        normalized = BotSpreadsheetService._normalize(value)
        if "tarjeta" in normalized or "card" in normalized:
            return "tarjeta"
        if "footer" in normalized or "pie" in normalized:
            return "footer"
        if "lateral" in normalized or "side" in normalized:
            return "lateral"
        return ""

    def _header_index(self, rows: list[tuple[Any, ...]]) -> int | None:
        best = None
        score = 0
        for index, row in enumerate(rows[:15]):
            current = sum(1 for value in row if self._normalize(self._text(value)) in self._all_aliases())
            if current > score:
                score, best = current, index
        return best if score else None

    def _mapping(self, headers: list[str]) -> dict[str, int]:
        mapping = {}
        for key, aliases in self.ALIASES.items():
            for index, header in enumerate(headers):
                normalized = self._normalize(header)
                if normalized in {self._normalize(alias) for alias in aliases}:
                    mapping[key] = index
                    break
        return mapping

    def _suggestions(self, sheets: list[dict[str, Any]]) -> list[str]:
        suggestions = []
        for sheet in sheets:
            missing = [key for key in ("utel_url",) if key not in sheet["mapping"]]
            if not missing:
                suggestions.append(f"{sheet['name']}: compatible para ejecutar el Bot.")
            else:
                suggestions.append(f"{sheet['name']}: faltan columnas sugeridas: {', '.join(missing)}.")
        return suggestions

    def _all_aliases(self) -> set[str]:
        return {self._normalize(alias) for aliases in self.ALIASES.values() for alias in aliases}

    @staticmethod
    def _text(value: Any) -> str:
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _normalize(value: Any) -> str:
        # El mapeo también contiene metadatos de ejecución (por ejemplo,
        # selected_row_number), que pueden ser enteros y no nombres de columna.
        value = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii").lower()
        return re.sub(r"\s+", " ", value).strip()
