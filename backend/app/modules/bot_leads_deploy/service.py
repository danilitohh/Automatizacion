"""Resolución de país/nivel y catálogo para Bot Leads Deploy.

Este archivo contiene únicamente ajustes propios de Leads Deploy. No modifica
el comportamiento del Bot de nuevos productos.
"""

from __future__ import annotations

from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from ...services.program_rotation_service import ProgramRotationService
from .spreadsheet_service import LeadsDeploySpreadsheetService as BaseLeadsDeploySpreadsheetService


class LeadsDeploySpreadsheetService(BaseLeadsDeploySpreadsheetService):
    """Capa final del parser/catálogo usada por la API de Leads Deploy."""

    GLOBAL_COUNTRY_HINTS = (
        (("filipinas", "philippines"), "Filipinas"),
        (("india",), "India"),
        (("indonesia",), "Indonesia"),
    )

    @classmethod
    def effective_country(cls, country: str, level: str, url: str) -> str:
        """Resuelve filas Global usando el país embebido en Nivel cuando existe.

        La matriz Leads Deploy agrupa algunos mercados internacionales bajo
        ``Global`` y distingue el país dentro de Nivel (por ejemplo,
        ``India Bachelor``). El servicio base solo resolvía Filipinas; por eso
        India/Indonesia podían terminar buscando programas para ``Global`` y
        dejar ``utel_url`` vacío.
        """

        if cls._normalize(country) != "global":
            return super().effective_country(country, level, url)

        source = cls._normalize(f"{level} {url}")
        for aliases, resolved_country in cls.GLOBAL_COUNTRY_HINTS:
            if any(alias in source for alias in aliases):
                return resolved_country

        return super().effective_country(country, level, url)

    @classmethod
    def deploy_navigation_plan(cls, raw_level: str, country: str) -> dict[str, str]:
        """Interpreta Bachelor/Master y la modalidad indicada por Nivel."""

        level = cls._normalize(raw_level)
        # La columna Nivel define la modalidad; retirar el sufijo permite
        # resolver también Doctorado, Diplomado y los niveles internacionales.
        modality_match = re.search(r"\b(hibrida|ejecutiva)\b", level)
        if modality_match:
            base_level = re.sub(r"\b(hibrida|ejecutiva)\b", "", level).strip()
            plan = cls.deploy_navigation_plan(base_level, country)
            plan["modality"] = "Hibrida" if modality_match.group(1) == "hibrida" else "Ejecutiva"
            plan["navigation_modality"] = f"Modalidad {plan['modality'].lower()}"
            return plan
        country_key = cls._catalog_key(country)
        international = country_key in {
            "filipinas",
            "india",
            "indonesia",
            "global",
        }

        if international and ("master" in level or "maestr" in level):
            return {
                "modality": "Online",
                "level": "Master's Degree",
                "navigation_modality": "",
                "navigation_level": "Master",
                "navigation_sublevel": "",
            }

        if international and ("bachelor" in level or "licenc" in level):
            return {
                "modality": "Online",
                "level": "Bachelor's Degree",
                "navigation_modality": "",
                "navigation_level": "Bachelor",
                "navigation_sublevel": "",
            }

        return super().deploy_navigation_plan(raw_level, country)

    @classmethod
    def _catalog_modality_from_level(cls, raw_level: str) -> str:
        """Obtiene modalidad desde NIVEL cuando el catálogo no tiene esa columna.

        El catálogo actual usa hojas por país y columnas NIVEL, PROGRAMA, URL.
        Por eso valores como ``Maestría Híbrida`` deben conservar la palabra
        Híbrida; reducirlos solo a ``master`` mezcla sus URLs con Maestría online.
        """

        level = cls._normalize(raw_level)
        if "hibr" in level:
            return "Hibrida"
        if "ejecut" in level:
            return "Ejecutiva"
        return "En linea"

    def catalog_programs(
        self,
        country: str,
        level: str,
        modality: str = "",
    ) -> list[dict[str, str]]:
        """Filtra el catálogo por país + nivel + modalidad real de Leads Deploy.

        Cuando el Excel de catálogo no tiene una columna MODALIDAD, la modalidad
        se infiere desde NIVEL. Esto evita que ``Maestría Híbrida`` comparta el
        mismo conjunto de candidatos con una ``Maestría`` normal.
        """

        if not self.catalog_path.is_file():
            return []

        country_key = self._catalog_key(country)
        level_key = self._catalog_level_key(level)
        modality_key = self._catalog_modality_key(modality)
        workbook = load_workbook(
            self.catalog_path,
            read_only=True,
            data_only=True,
        )
        matches: list[dict[str, str]] = []

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, row in enumerate(values[:20])
                    if self._catalog_header_row(row)
                ),
                None,
            )
            if header_index is None:
                continue

            headers = [
                self._normalize(self._text(value))
                for value in values[header_index]
            ]
            indexes = {
                "country": self._catalog_index(headers, ("pais", "country")),
                "modality": self._catalog_index(headers, ("modalidad", "modality")),
                "level": self._catalog_index(headers, ("nivel", "level")),
                "program": self._catalog_index(headers, ("programa", "program")),
                "url": self._catalog_index(headers, ("url del programa", "program url", "url")),
            }
            if indexes["program"] is None or indexes["url"] is None:
                continue

            last_country = worksheet.title if indexes["country"] is None else ""
            last_level = ""
            last_modality = ""

            for row in values[header_index + 1 :]:
                row_country = self._cell(row, indexes["country"])
                row_level = self._cell(row, indexes["level"])
                row_modality = self._cell(row, indexes["modality"])

                if row_country:
                    last_country = row_country
                else:
                    row_country = last_country

                if row_level:
                    last_level = row_level
                else:
                    row_level = last_level

                if indexes["modality"] is None:
                    # Catálogo nuevo: la modalidad forma parte del texto NIVEL.
                    row_modality = self._catalog_modality_from_level(row_level)
                elif row_modality:
                    last_modality = row_modality
                else:
                    row_modality = last_modality

                program = self._cell(row, indexes["program"])
                url = self._cell(row, indexes["url"])
                if not program or not url:
                    continue
                if self._catalog_key(row_country) != country_key:
                    continue
                if not self._catalog_level_matches(
                    level_key,
                    self._catalog_level_key(row_level),
                ):
                    continue
                if modality_key and not self._catalog_modality_matches(
                    modality_key,
                    self._catalog_modality_key(row_modality),
                ):
                    continue
                matches.append({"text": program, "url": url})

        return list(
            {
                (item["text"], item["url"]): item
                for item in matches
            }.values()
        )

    def _rotate_catalog_candidate(
        self,
        candidates: list[dict[str, str]],
        country: str,
        level: str,
        modality: str,
        database_path: Path,
    ) -> dict[str, str]:
        return ProgramRotationService(database_path).choose(
            [
                "leads_deploy_catalog",
                self._catalog_key(country),
                self._catalog_level_key(level),
                self._catalog_modality_key(modality),
            ],
            candidates,
        )

    def choose_catalog_program(
        self,
        country: str,
        level: str,
        modality: str,
        database_path: Path,
    ) -> dict[str, str]:
        """Garantiza que Híbrida/Ejecutiva nunca caigan a un programa online.

        Primero se busca una coincidencia estricta. Para Híbrida y Ejecutiva no
        existe fallback sin modalidad: si el catálogo no contiene esa combinación
        es preferible fallar con un mensaje claro a abrir una URL equivocada.
        """

        candidates = self.catalog_programs(country, level, modality)
        if candidates:
            return self._rotate_catalog_candidate(
                candidates,
                country,
                level,
                modality,
                database_path,
            )

        navigation = self.deploy_navigation_plan(level, country)
        alternate_level = navigation.get("level", "")
        alternate_modality = navigation.get("modality", "") or modality
        changed_level = (
            self._catalog_level_key(alternate_level)
            != self._catalog_level_key(level)
        )

        if changed_level:
            candidates = self.catalog_programs(
                country,
                alternate_level,
                alternate_modality,
            )
            if candidates:
                return self._rotate_catalog_candidate(
                    candidates,
                    country,
                    alternate_level,
                    alternate_modality,
                    database_path,
                )

        requested_modality = self._catalog_modality_key(modality)
        strict_modalities = {"hibrida", "ejecutiva"}

        # Para modalidades no estrictas se conserva el fallback histórico por
        # compatibilidad. Híbrida/Ejecutiva jamás deben mezclarse con En línea.
        if modality and requested_modality not in strict_modalities:
            candidates = self.catalog_programs(country, level, "")
            if candidates:
                return self._rotate_catalog_candidate(
                    candidates,
                    country,
                    level,
                    "",
                    database_path,
                )

        raise ValueError(
            "Leads Deploy no encontró un programa con URL de la modalidad solicitada "
            f"para País='{country}', Nivel='{level}', Modalidad='{modality}'. "
            "Revisa esa combinación en backend/data/Programas_UTEL_Todos_los_Paises.xlsx."
        )
