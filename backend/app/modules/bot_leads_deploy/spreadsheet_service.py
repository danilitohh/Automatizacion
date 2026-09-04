"""Servicio de Excel y catálogo exclusivo de Bot Leads Deploy.

Este módulo acepta la matriz Leads Deploy sin depender de la columna
"Activo de Test". Cada caso se define por País + Nivel + Location y el
programa/URL se obtiene del catálogo interno.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ...services.bot_spreadsheet_service import BotSpreadsheetService as BaseBotSpreadsheetService
from ...services.program_rotation_service import ProgramRotationService


class LeadsDeploySpreadsheetService(BaseBotSpreadsheetService):
    """Parser y resolvedor de catálogo específico de Leads Deploy."""

    COUNTRY_ALIASES = {
        "mexico": "mexico",
        "ecuador": "ecuador",
        "colombia": "colombia",
        "peru": "peru",
        "argentina": "argentina",
        "usa": "usa",
        "united states": "usa",
        "estados unidos": "usa",
        "bolivia": "bolivia",
        "chile": "chile",
        "paraguay": "paraguay",
        "dominicana": "dominicana",
        "republica dominicana": "dominicana",
        "dominican republic": "dominicana",
        "guatemala": "guatemala",
        "panama": "panama",
        "el salvador": "el salvador",
        "filipinas": "filipinas",
        "philippines": "filipinas",
        "india": "india",
        "indonesia": "indonesia",
        "global": "global",
    }

    @classmethod
    def _catalog_key(cls, value: str) -> str:
        """Normaliza aliases de país tanto del Excel Deploy como del catálogo."""

        key = cls._normalize(value)
        return cls.COUNTRY_ALIASES.get(key, key)

    def catalog_programs(
        self,
        country: str,
        level: str,
        modality: str = "",
    ) -> list[dict[str, str]]:
        """Busca programas tolerando celdas agrupadas/vacías del catálogo.

        El catálogo suele agrupar país, nivel o modalidad visualmente. Para
        Leads Deploy heredamos el último valor no vacío de esas columnas, sin
        cambiar el comportamiento del Bot de nuevos productos.
        """

        if not self.catalog_path.is_file():
            return []

        country_key = self._catalog_key(country)
        level_key = self._catalog_level_key(level)
        modality_key = self._catalog_modality_key(modality)
        workbook = load_workbook(
            self.catalog_path,
            read_only=True,
            data_only=True,
        )
        matches: list[dict[str, str]] = []

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = next(
                (
                    index
                    for index, row in enumerate(values[:20])
                    if self._catalog_header_row(row)
                ),
                None,
            )
            if header_index is None:
                continue

            headers = [
                self._normalize(self._text(value))
                for value in values[header_index]
            ]
            indexes = {
                "country": self._catalog_index(headers, ("pais", "country")),
                "modality": self._catalog_index(headers, ("modalidad", "modality")),
                "level": self._catalog_index(headers, ("nivel", "level")),
                "program": self._catalog_index(headers, ("programa", "program")),
                "url": self._catalog_index(headers, ("url del programa", "program url")),
            }
            if indexes["program"] is None or indexes["url"] is None:
                continue

            last_country = ""
            last_level = ""
            last_modality = ""

            for row in values[header_index + 1 :]:
                row_country = self._cell(row, indexes["country"])
                row_level = self._cell(row, indexes["level"])
                row_modality = self._cell(row, indexes["modality"])

                if row_country:
                    last_country = row_country
                else:
                    row_country = last_country

                if row_level:
                    last_level = row_level
                else:
                    row_level = last_level

                if row_modality:
                    last_modality = row_modality
                else:
                    row_modality = last_modality

                program = self._cell(row, indexes["program"])
                url = self._cell(row, indexes["url"])
                if not program or not url:
                    continue
                if self._catalog_key(row_country) != country_key:
                    continue
                if not self._catalog_level_matches(
                    level_key,
                    self._catalog_level_key(row_level),
                ):
                    continue
                if modality_key and not self._catalog_modality_matches(
                    modality_key,
                    self._catalog_modality_key(row_modality),
                ):
                    continue
                matches.append({"text": program, "url": url})

        return list(
            {
                (item["text"], item["url"]): item
                for item in matches
            }.values()
        )

    def choose_catalog_program(
        self,
        country: str,
        level: str,
        modality: str,
        database_path: Path,
    ) -> dict[str, str] | None:
        """Resuelve un programa antes de validar UtelQaConfig.

        Primero exige país+nivel+modalidad. Si el catálogo usa una etiqueta de
        modalidad diferente, relaja únicamente la modalidad; nunca cambia de
        país ni de nivel. Así evitamos dejar ``utel_url=''`` y obtener un error
        genérico de Pydantic.
        """

        candidates = self.catalog_programs(country, level, modality)
        scope_modality = modality

        if not candidates and modality:
            candidates = self.catalog_programs(country, level, "")
            scope_modality = ""

        if not candidates:
            return None

        return ProgramRotationService(database_path).choose(
            [
                "leads_deploy_catalog",
                self._catalog_key(country),
                self._catalog_level_key(level),
                self._catalog_modality_key(scope_modality),
            ],
            candidates,
        )

    def preview(self, content: bytes, filename: str) -> dict[str, Any]:
        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
        sheets: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = self._header_index(values)
            if header_index is None:
                continue

            headers = [self._text(value) for value in values[header_index]]
            detected_mapping = self._mapping(headers)

            if not all(
                key in detected_mapping
                for key in ("country", "level", "form_type")
            ):
                continue

            detected_mapping.pop("utel_url", None)
            normalized_rows: list[dict[str, Any]] = []
            last_country = ""

            for row_number, row_values in enumerate(
                values[header_index + 1 :],
                header_index + 2,
            ):
                if not any(self._text(value) for value in row_values):
                    continue

                item = {
                    key: self._text(row_values[index])
                    if index < len(row_values)
                    else ""
                    for key, index in detected_mapping.items()
                }

                current_country = item.get("country", "")
                if current_country:
                    last_country = current_country
                else:
                    current_country = last_country

                level = item.get("level", "")
                form_type = self._normalize_form_type(
                    item.get("form_type", "")
                )
                if not current_country or not level or not form_type:
                    continue

                item.update(
                    {
                        "country": current_country,
                        "level": level,
                        "form_type": form_type,
                        "utel_url": "",
                        "workflow_mode": "form_validation",
                    }
                )
                normalized_rows.append(
                    {"row_number": row_number, **item}
                )

            if normalized_rows:
                sheets.append(
                    {
                        "name": worksheet.title,
                        "headers": headers,
                        "mapping": {
                            key: headers[index]
                            for key, index in detected_mapping.items()
                        },
                        "rows": normalized_rows[:200],
                    }
                )

        return {
            "filename": filename,
            "sheets": sheets,
            "suggestions": [
                (
                    f"{sheet['name']}: {len(sheet['rows'])} casos Leads Deploy "
                    "detectados. Activo de Test se ignora; las URLs se toman "
                    "del catálogo interno."
                )
                for sheet in sheets
            ],
        }

    def rows_for_mapping(
        self,
        content: bytes,
        mapping: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Lee casos Deploy sin exigir URL de programa en el Excel."""

        workbook = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
        selected = {
            key: self._normalize(value)
            for key, value in mapping.items()
            if isinstance(value, str) and value
        }

        required_mapping = ("country", "level", "form_type")
        if not all(selected.get(key) for key in required_mapping):
            return []

        rows: list[dict[str, Any]] = []

        for worksheet in workbook.worksheets:
            values = list(worksheet.iter_rows(values_only=True))
            header_index = self._header_index(values)
            if header_index is None:
                continue

            headers = [self._text(value) for value in values[header_index]]
            indexes = {
                key: next(
                    (
                        index
                        for index, header in enumerate(headers)
                        if self._normalize(header) == selected_value
                    ),
                    None,
                )
                for key, selected_value in selected.items()
            }
            if any(
                indexes.get(key) is None
                for key in required_mapping
            ):
                continue

            last_country = ""
            for row_number, row_values in enumerate(
                values[header_index + 1 :],
                header_index + 2,
            ):
                level = self._cell(row_values, indexes.get("level"))
                form_type = self._normalize_form_type(
                    self._cell(row_values, indexes.get("form_type"))
                )
                current_country = self._cell(
                    row_values,
                    indexes.get("country"),
                )

                if current_country:
                    last_country = current_country
                else:
                    current_country = last_country

                if not level or not form_type or not current_country:
                    continue

                program = self._cell(
                    row_values,
                    indexes.get("program_name"),
                )
                rows.append(
                    {
                        "sheet": worksheet.title,
                        "row_number": row_number,
                        "program_name": program,
                        "level": level,
                        "modality": self._cell(
                            row_values,
                            indexes.get("modality"),
                        ),
                        "country": current_country,
                        "form_type": form_type,
                        "inconcert_url": self._cell(
                            row_values,
                            indexes.get("inconcert_url"),
                        ),
                        "lead_origin_url": self._cell(
                            row_values,
                            indexes.get("lead_origin_url"),
                        ),
                        "utel_url": "",
                        "workflow_mode": "form_validation",
                        "test_case": level or program,
                    }
                )

        return rows
