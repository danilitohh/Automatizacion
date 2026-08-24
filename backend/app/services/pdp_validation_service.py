"""Validación de contenido entre una matriz Excel, un DOCX y páginas PDP."""

from __future__ import annotations

import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from difflib import SequenceMatcher
from io import BytesIO
from pathlib import Path
from time import perf_counter
from typing import Any, Iterable

from docx import Document
from docx.document import Document as DocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook

from ..config.settings import Settings
from .logging_service import get_logger


MAX_FILE_SIZE = 12 * 1024 * 1024
MAX_PROGRAMS = 100
STOP_WORDS = {
    "a", "al", "con", "de", "del", "el", "en", "es", "la", "las", "lo",
    "los", "para", "por", "que", "se", "su", "una", "un", "y", "o",
}
SECTION_ALIASES = {
    "description": ("descripcion", "acerca", "sobre el programa", "presentacion"),
    "subjects": ("asignaturas", "materias", "plan de estudios", "mapa curricular"),
    "faqs": ("preguntas frecuentes", "preguntas y respuestas", "faq", "faqs"),
}


@dataclass
class ProgramReference:
    """Fila utilizable de la matriz Excel."""

    name: str
    url: str
    row_number: int


@dataclass
class ReferenceContent:
    """Contenido de referencia de un programa extraído del DOCX."""

    title: str = ""
    description: list[str] = field(default_factory=list)
    subjects: list[str] = field(default_factory=list)
    faqs: list[str] = field(default_factory=list)
    found: bool = False


class PdpValidationService:
    """Orquesta lectura de archivos y comparación de contenido web visible."""

    def __init__(self, settings: Settings):
        self.settings = settings
        self.logger = get_logger()

    async def validate(self, excel_bytes: bytes, docx_bytes: bytes) -> dict[str, Any]:
        """Valida cada PDP de la matriz contra la información del documento."""

        started_at = datetime.now().isoformat(timespec="seconds")
        started_timer = perf_counter()
        programs, excel_info = self._read_excel(excel_bytes)
        blocks = self._read_docx(docx_bytes)
        references = self._map_document_to_programs(programs, blocks)
        results = await self._validate_pages(programs, references)
        summary = self._build_summary(results)
        finished_at = datetime.now().isoformat(timespec="seconds")
        report = {
            "status": "PASS" if summary["failed"] == 0 and summary["errors"] == 0 else "WARNING",
            "summary": summary,
            "excel": excel_info,
            "programs": results,
            "started_at": started_at,
            "finished_at": finished_at,
            "duration_seconds": round(perf_counter() - started_timer, 2),
        }
        report["report_file"] = self._save_report(report)
        return report

    def _read_excel(self, content: bytes) -> tuple[list[ProgramReference], dict[str, Any]]:
        """Detecta columnas de programa y URL sin exigir un nombre exacto."""

        if not content:
            raise ValueError("El archivo Excel está vacío.")
        if len(content) > MAX_FILE_SIZE:
            raise ValueError("El Excel supera el límite de 12 MB.")

        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as error:
            raise ValueError("No fue posible leer el Excel. Usa un archivo .xlsx válido.") from error

        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as error:
            raise ValueError("El Excel no contiene encabezados ni filas.") from error

        headers = [self._normalize_header(value) for value in raw_headers]
        program_index = self._find_header(headers, "program")
        url_index = self._find_header(headers, "url")
        if program_index is None or url_index is None:
            available = ", ".join(str(value) for value in raw_headers if value)
            raise ValueError(
                "No se detectaron las columnas requeridas. Incluye una columna de programa "
                "(Programa, Carrera o Nombre) y otra de URL (URL, Link, Enlace o PDP). "
                f"Encabezados encontrados: {available or 'ninguno'}."
            )

        programs: list[ProgramReference] = []
        skipped = 0
        for row_number, row in enumerate(rows, start=2):
            name = self._clean_text(row[program_index] if program_index < len(row) else "")
            url = self._clean_text(row[url_index] if url_index < len(row) else "")
            if not name and not url:
                continue
            if not name or not self._is_http_url(url):
                skipped += 1
                continue
            programs.append(ProgramReference(name=name, url=url, row_number=row_number))
            if len(programs) >= MAX_PROGRAMS:
                break

        workbook.close()
        if not programs:
            raise ValueError("No hay filas válidas: cada una necesita nombre de programa y URL http(s).")
        return programs, {
            "sheet": sheet.title,
            "program_column": str(raw_headers[program_index]),
            "url_column": str(raw_headers[url_index]),
            "programs_loaded": len(programs),
            "rows_skipped": skipped,
            "limited_to": MAX_PROGRAMS if len(programs) == MAX_PROGRAMS else None,
        }

    def _read_docx(self, content: bytes) -> list[dict[str, str]]:
        """Lee párrafos y tablas respetando el orden en el que aparecen."""

        if not content:
            raise ValueError("El documento DOCX está vacío.")
        if len(content) > MAX_FILE_SIZE:
            raise ValueError("El DOCX supera el límite de 12 MB.")
        try:
            document = Document(BytesIO(content))
        except Exception as error:
            raise ValueError("No fue posible leer el DOCX. Usa un archivo .docx válido.") from error

        blocks: list[dict[str, str]] = []
        for block in self._iter_docx_blocks(document):
            if isinstance(block, Paragraph):
                text = self._clean_text(block.text)
                if text:
                    blocks.append({"kind": "paragraph", "text": text})
            elif isinstance(block, Table):
                for row in block.rows:
                    values = [self._clean_text(cell.text) for cell in row.cells]
                    values = list(dict.fromkeys(value for value in values if value))
                    if values:
                        blocks.append({"kind": "table", "text": " | ".join(values)})
        if not blocks:
            raise ValueError("No se encontró texto utilizable dentro del DOCX.")
        return blocks

    async def _validate_pages(
        self,
        programs: list[ProgramReference],
        references: dict[str, ReferenceContent],
    ) -> list[dict[str, Any]]:
        """Abre cada URL y compara el texto de su PDP con el documento."""

        try:
            from playwright.async_api import async_playwright
        except ImportError as error:
            raise RuntimeError("Playwright no está instalado para revisar las páginas PDP.") from error

        results: list[dict[str, Any]] = []
        async with async_playwright() as playwright:
            browser = await playwright.chromium.launch(headless=True)
            context = await browser.new_context(viewport={"width": 1440, "height": 1000})
            try:
                for program in programs:
                    reference = references.get(self._normalize(program.name), ReferenceContent())
                    results.append(await self._validate_program(context, program, reference))
            finally:
                await context.close()
                await browser.close()
        return results

    async def _validate_program(
        self,
        context: Any,
        program: ProgramReference,
        reference: ReferenceContent,
    ) -> dict[str, Any]:
        """Construye el reporte de una PDP sin exponer errores técnicos crudos."""

        base = {
            "program": program.name,
            "url": program.url,
            "excel_row": program.row_number,
            "reference_found": reference.found,
        }
        if not reference.found:
            return {
                **base,
                "status": "WARNING",
                "message": "No se encontró una sección del DOCX asociada a este programa.",
                "title": self._not_available(),
                "description": self._not_available(),
                "subjects": self._not_available(),
                "faqs": self._not_available(),
            }

        page = await context.new_page()
        page.set_default_timeout(20_000)
        try:
            await page.goto(program.url, wait_until="domcontentloaded", timeout=35_000)
            await page.wait_for_timeout(900)
            web_content = await page.evaluate(
                """() => ({
                    title: document.querySelector('h1')?.innerText || document.title || '',
                    headings: [...document.querySelectorAll('h1,h2,h3,h4')]
                        .map((heading) => heading.innerText || '')
                        .filter(Boolean),
                    text: document.body?.innerText || document.body?.textContent || ''
                })"""
            )
            title = self._compare_title(reference.title or program.name, web_content["title"], web_content["headings"])
            description = self._compare_section(reference.description, web_content["text"], "Descripción")
            subjects = self._compare_items(reference.subjects, web_content["text"], "Asignaturas")
            faqs = self._compare_items(reference.faqs, web_content["text"], "Preguntas frecuentes")
            states = [title["status"], description["status"], subjects["status"], faqs["status"]]
            status = "PASS" if all(state == "PASS" for state in states if state != "NO_DISPONIBLE") else "WARNING"
            return {
                **base,
                "status": status,
                "message": "PDP comparada correctamente.",
                "title": title,
                "description": description,
                "subjects": subjects,
                "faqs": faqs,
            }
        except Exception as error:
            self.logger.warning("No se pudo validar PDP %s: %s", program.url, error)
            return {
                **base,
                "status": "ERROR",
                "message": self._friendly_error(error),
                "title": self._not_available(),
                "description": self._not_available(),
                "subjects": self._not_available(),
                "faqs": self._not_available(),
            }
        finally:
            await page.close()

    def _map_document_to_programs(
        self,
        programs: list[ProgramReference],
        blocks: list[dict[str, str]],
    ) -> dict[str, ReferenceContent]:
        """Separa un único DOCX por programa usando sus títulos como límites."""

        locations: list[tuple[int, ProgramReference]] = []
        for program in programs:
            match = self._find_program_heading(program.name, blocks)
            if match is not None:
                locations.append((match, program))
        locations.sort(key=lambda item: item[0])

        references: dict[str, ReferenceContent] = {}
        for position, (start, program) in enumerate(locations):
            end = locations[position + 1][0] if position + 1 < len(locations) else len(blocks)
            references[self._normalize(program.name)] = self._extract_reference(program.name, blocks[start:end])

        if len(programs) == 1 and not references:
            references[self._normalize(programs[0].name)] = self._extract_reference(programs[0].name, blocks)
        return references

    def _extract_reference(self, program_name: str, blocks: Iterable[dict[str, str]]) -> ReferenceContent:
        """Clasifica el bloque de un programa en descripción, materias y FAQ."""

        reference = ReferenceContent(title=program_name, found=True)
        current_section = "description"
        for block in blocks:
            text = self._clean_text(block["text"])
            if not text:
                continue
            section = self._section_for_heading(text)
            if section:
                current_section = section
                continue
            if self._normalize(text) == self._normalize(program_name):
                reference.title = text
                continue
            if current_section == "description":
                reference.description.append(text)
            elif current_section == "subjects":
                reference.subjects.extend(self._split_items(text, is_table=block["kind"] == "table"))
            elif current_section == "faqs":
                reference.faqs.extend(self._split_items(text, is_table=False))

        reference.description = self._unique_items(reference.description)
        reference.subjects = self._unique_items(reference.subjects)
        reference.faqs = self._unique_items(reference.faqs)
        return reference

    @staticmethod
    def _iter_docx_blocks(document: DocumentType) -> Iterable[Paragraph | Table]:
        parent = document.element.body
        for child in parent.iterchildren():
            if isinstance(child, CT_P):
                yield Paragraph(child, document)
            elif isinstance(child, CT_Tbl):
                yield Table(child, document)

    def _find_program_heading(self, program_name: str, blocks: list[dict[str, str]]) -> int | None:
        expected = self._normalize(program_name)
        best_index: int | None = None
        best_score = 0.0
        for index, block in enumerate(blocks):
            if block["kind"] != "paragraph":
                continue
            candidate = self._normalize(block["text"])
            if not candidate:
                continue
            score = self._similarity(expected, candidate)
            if score > best_score:
                best_index, best_score = index, score
        return best_index if best_score >= 0.72 else None

    def _compare_title(self, expected: str, actual: str, headings: list[str]) -> dict[str, Any]:
        best_actual = actual or (headings[0] if headings else "")
        score = max(
            [self._similarity(expected, best_actual)] + [self._similarity(expected, heading) for heading in headings]
        )
        return {
            "status": "PASS" if score >= 0.72 else "WARNING",
            "expected": expected,
            "found": best_actual,
            "coverage": round(score * 100),
        }

    def _compare_section(self, expected_items: list[str], web_text: str, label: str) -> dict[str, Any]:
        if not expected_items:
            return self._not_available()
        expected = " ".join(expected_items)
        coverage = self._token_coverage(expected, web_text)
        return {
            "status": self._coverage_status(coverage),
            "expected_items": len(expected_items),
            "coverage": round(coverage * 100),
            "missing": [] if coverage >= 0.65 else [self._shorten(expected)],
            "label": label,
        }

    def _compare_items(self, expected_items: list[str], web_text: str, label: str) -> dict[str, Any]:
        if not expected_items:
            return self._not_available()
        found: list[str] = []
        missing: list[str] = []
        normalized_web = self._normalize(web_text)
        for item in expected_items:
            normalized_item = self._normalize(item)
            coverage = self._token_coverage(item, web_text)
            if normalized_item and (normalized_item in normalized_web or coverage >= 0.72):
                found.append(item)
            else:
                missing.append(item)
        coverage = len(found) / len(expected_items)
        return {
            "status": self._coverage_status(coverage),
            "expected_items": len(expected_items),
            "found_items": len(found),
            "coverage": round(coverage * 100),
            "missing": [self._shorten(item) for item in missing[:12]],
            "label": label,
        }

    def _build_summary(self, results: list[dict[str, Any]]) -> dict[str, int]:
        status_counter = Counter(result["status"] for result in results)
        section_statuses = [
            section["status"]
            for result in results
            for section in (result["title"], result["description"], result["subjects"], result["faqs"])
            if section["status"] != "NO_DISPONIBLE"
        ]
        return {
            "programs": len(results),
            "passed": status_counter["PASS"],
            "warnings": status_counter["WARNING"],
            "errors": status_counter["ERROR"],
            "failed": sum(status in {"FAIL", "WARNING"} for status in section_statuses),
            "sections_checked": len(section_statuses),
        }

    def _save_report(self, report: dict[str, Any]) -> str:
        directory = self.settings.storage_dir / "reports" / "pdp"
        directory.mkdir(parents=True, exist_ok=True)
        file_path = directory / f"pdp-validation-{datetime.now():%Y%m%d-%H%M%S}.json"
        file_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        return file_path.relative_to(self.settings.storage_dir.parent).as_posix()

    @staticmethod
    def _find_header(headers: list[str], field: str) -> int | None:
        aliases = {
            "program": ("programa", "carrera", "nombre", "nombreprograma", "nombrecarrera", "producto"),
            "url": ("url", "urlpdp", "link", "enlace", "pdp", "paginaproducto", "paginadeproducto"),
        }[field]
        for index, header in enumerate(headers):
            if header in aliases or any(alias in header for alias in aliases):
                return index
        return None

    def _section_for_heading(self, text: str) -> str | None:
        normalized = self._normalize(text)
        for section, aliases in SECTION_ALIASES.items():
            if any(alias in normalized for alias in aliases):
                return section
        return None

    @staticmethod
    def _split_items(text: str, is_table: bool) -> list[str]:
        if is_table:
            values = [item.strip() for item in text.split("|")]
            table_headers = {"periodo", "cuatrimestre", "semestre", "asignatura", "materia", "clave", "creditos"}
            return [
                item for item in values
                if len(item) > 2 and PdpValidationService._normalize(item) not in table_headers
            ]
        values = re.split(r"\n|•|\s+[-–]\s+", text)
        return [value.strip() for value in values if len(value.strip()) > 2]

    @staticmethod
    def _unique_items(items: list[str]) -> list[str]:
        seen: set[str] = set()
        result: list[str] = []
        for item in items:
            normalized = PdpValidationService._normalize(item)
            if normalized and normalized not in seen:
                seen.add(normalized)
                result.append(item)
        return result

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return PdpValidationService._normalize(str(value or "")).replace(" ", "")

    @staticmethod
    def _normalize(value: str) -> str:
        value = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"\s+", " ", re.sub(r"[^a-zA-Z0-9]+", " ", value.lower())).strip()

    @staticmethod
    def _clean_text(value: Any) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    @staticmethod
    def _is_http_url(value: str) -> bool:
        return value.lower().startswith(("http://", "https://"))

    def _similarity(self, expected: str, actual: str) -> float:
        expected_normalized = self._normalize(expected)
        actual_normalized = self._normalize(actual)
        if not expected_normalized or not actual_normalized:
            return 0.0
        if expected_normalized in actual_normalized or actual_normalized in expected_normalized:
            return 1.0
        return max(
            SequenceMatcher(None, expected_normalized, actual_normalized).ratio(),
            self._token_coverage(expected_normalized, actual_normalized),
        )

    def _token_coverage(self, expected: str, actual: str) -> float:
        expected_tokens = {token for token in self._normalize(expected).split() if len(token) > 2 and token not in STOP_WORDS}
        actual_tokens = set(self._normalize(actual).split())
        if not expected_tokens:
            return 1.0
        return len(expected_tokens & actual_tokens) / len(expected_tokens)

    @staticmethod
    def _coverage_status(coverage: float) -> str:
        if coverage >= 0.72:
            return "PASS"
        if coverage >= 0.4:
            return "WARNING"
        return "FAIL"

    @staticmethod
    def _not_available() -> dict[str, Any]:
        return {"status": "NO_DISPONIBLE", "coverage": 0, "expected_items": 0, "missing": []}

    @staticmethod
    def _shorten(value: str, length: int = 150) -> str:
        return value if len(value) <= length else f"{value[:length - 1].rstrip()}…"

    @staticmethod
    def _friendly_error(error: Exception) -> str:
        message = re.sub(r"\s+", " ", str(error)).strip()
        return message[:240] or "No se pudo abrir la página PDP."
