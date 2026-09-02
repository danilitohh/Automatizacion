"""Genera una copia del Excel sin modificar celdas combinadas ni datos de entrada."""

import io
from openpyxl import load_workbook

from .bot_spreadsheet_service import BotSpreadsheetService


class BotReportService:
    def build(self, content: bytes, mapping: dict, results: list):
        workbook = load_workbook(io.BytesIO(content))
        service = BotSpreadsheetService()
        requested = mapping.get("lead_url") or "URL LEAD"
        inputs = {service._normalize(mapping.get(key, "")) for key in
                  ("country", "level", "program_name", "modality", "utel_url", "inconcert_url", "form_type")}
        reserved = {service._normalize(label) for label in ("RESULTADO FORMULARIO", "DETALLE BOT", "EMAIL LEAD PRUEBA", "PROGRAMA SELECCIONADO BOT")}
        if service._normalize(requested) in (inputs - {service._normalize(mapping.get("inconcert_url", ""))}) | reserved:
            requested = "URL LEAD"
        for sheet in workbook.worksheets:
            items = [item for item in results if item["row"]["sheet"] == sheet.title]
            if not items:
                continue
            header = service._header_index(list(sheet.iter_rows(values_only=True)))
            if header is None:
                continue
            header += 1

            def existing_column(label):
                if not service._normalize(label or ""):
                    return None
                for cell in sheet[header]:
                    if service._normalize(service._text(cell.value)) == service._normalize(label):
                        if not any(merged.min_col <= cell.column <= merged.max_col for merged in sheet.merged_cells.ranges):
                            return cell.column
                return None

            def column(label):
                current = existing_column(label)
                if current is not None:
                    return current
                # max_column incluye celdas combinadas y columnas sin título.
                index = sheet.max_column + 1
                sheet.cell(header, index, label)
                return index

            def is_safe_empty_output_column(index):
                if index is None:
                    return False
                for item in items:
                    row = item["row"]["row_number"]
                    if sheet.cell(row, index).value not in (None, ""):
                        return False
                    if any(
                        merged.min_row <= row <= merged.max_row
                        and merged.min_col <= index <= merged.max_col
                        for merged in sheet.merged_cells.ranges
                    ):
                        return False
                return True

            status_col = column("RESULTADO FORMULARIO")
            detail_col = column("DETALLE BOT")
            requested_col = existing_column(requested)
            inconcert_col = existing_column(mapping.get("inconcert_url", ""))
            output_label = requested
            if requested_col == inconcert_col and not is_safe_empty_output_column(inconcert_col):
                output_label = "URL LEAD"
                requested_col = existing_column(output_label)
            # Leads Deploy ya incluye "inconcert/balanceador" como columna de
            # salida vacía. Úsala en lugar de crear "URL LEAD" al final, pero
            # nunca sobrescribas URLs de acceso ni celdas combinadas existentes.
            if (
                is_safe_empty_output_column(inconcert_col)
                and (
                    requested_col is None
                    and service._normalize(requested) == service._normalize("URL LEAD")
                    or requested_col == inconcert_col
                )
            ):
                link_col = inconcert_col
            else:
                link_col = column(output_label)
            email_col = column("EMAIL LEAD PRUEBA")
            program_col = column("PROGRAMA SELECCIONADO BOT")
            for item in items:
                result = item["result"]
                row = item["row"]["row_number"]
                failure = next((s for s in result.get("stages", []) if s["status"] == "FAIL"), None)
                dry_run = result.get("dry_run", False)
                link = None if dry_run else result.get("lead_url")
                status = "ERROR"
                if result["status"] == "PASS":
                    status = "DRY RUN - NO ENVIADO" if dry_run else "EXITOSO" if link else "SIN LINK VERIFICADO"
                elif link and failure and failure.get("stage") == "inconcert_manage":
                    status = "LEAD LOCALIZADO - VALIDACION PENDIENTE"
                sheet.cell(row, status_col).value = status
                detail = failure["message"] if failure else result.get("summary", "")
                if status == "LEAD LOCALIZADO - VALIDACION PENDIENTE":
                    detail = f"Lead localizado y enlace guardado. Faltó validar visualmente el email en InConcert: {detail}"
                sheet.cell(row, detail_col).value = detail[:32767]
                cell = sheet.cell(row, link_col)
                cell.value = link
                cell.hyperlink = link
                if link:
                    cell.style = "Hyperlink"
                sheet.cell(row, email_col).value = result.get("lead_email")
                sheet.cell(row, program_col).value = result.get("selected_program_name")
        return workbook
