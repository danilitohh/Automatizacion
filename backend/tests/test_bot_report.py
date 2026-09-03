from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from backend.app.services.bot_report_service import BotReportService


@pytest.mark.parametrize("dry_run,failed", [(False, False), (True, False), (False, True)])
def test_report_preserves_merges_and_source_crm(dry_run, failed):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nivel", "URL", "Location", "Locale", "CRM"])
    sheet.append(["Licenciatura", "https://utel.test", "footer", "Mexico", "https://crm.test/login"])
    sheet.merge_cells("E2:X2")
    data = BytesIO()
    workbook.save(data)
    link = "https://crm.test/mas/contact/people/view/123"
    result = {"row": {"sheet": "Sheet", "row_number": 2},
              "result": {"status": "FAIL" if failed else "PASS", "dry_run": dry_run,
                         "lead_url": None if failed else link, "lead_email": "qa@example.test",
                         "summary": "Prueba", "stages": [{"status": "FAIL", "message": "CRM no disponible"}] if failed else []}}
    output = BotReportService().build(data.getvalue(), {"inconcert_url": "CRM", "lead_url": "CRM"}, [result])
    saved = BytesIO()
    output.save(saved)
    sheet = load_workbook(BytesIO(saved.getvalue())).active
    assert str(next(iter(sheet.merged_cells.ranges))) == "E2:X2"
    assert sheet["E2"].value == "https://crm.test/login"
    columns = {cell.value: cell.column for cell in sheet[1] if cell.value}
    assert columns["URL LEAD"] > 24
    cell = sheet.cell(2, columns["URL LEAD"])
    assert cell.value == (None if dry_run or failed else link)
    if not dry_run and not failed:
        assert cell.hyperlink.target == link
    assert sheet.cell(2, columns["RESULTADO FORMULARIO"]).value == (
        "ERROR" if failed else "DRY RUN - NO ENVIADO" if dry_run else "EXITOSO")


def test_dry_run_clears_previous_lead_link_only_for_processed_rows():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nivel", "URL", "URL LEAD"])
    for _ in range(2):
        sheet.append(["Licenciatura", "https://utel.test", "https://crm.test/old"])
    sheet["C2"].hyperlink = "https://crm.test/old"
    data = BytesIO()
    workbook.save(data)
    result = {"row": {"sheet": "Sheet", "row_number": 2},
              "result": {"status": "PASS", "dry_run": True, "stages": [], "summary": "No enviado"}}
    sheet = BotReportService().build(data.getvalue(), {}, [result]).active
    assert sheet["C2"].value is None and sheet["C2"].hyperlink is None
    assert sheet["C3"].value == "https://crm.test/old"


def test_report_writes_lead_into_blank_inconcert_balanceador_column():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nivel", "URL", "Location", "Locale", "inconcert/balanceador"])
    sheet.append(["Licenciatura", "https://utel.test", "footer", "Mexico", None])
    data = BytesIO()
    workbook.save(data)
    link = "https://mas-utel.inconcertcc.com/mas/contact/people/view/123"
    result = {
        "row": {"sheet": "Sheet", "row_number": 2},
        "result": {
            "status": "PASS",
            "dry_run": False,
            "lead_url": link,
            "lead_email": "qa@example.test",
            "summary": "Verificado",
            "stages": [],
        },
    }

    output = BotReportService().build(
        data.getvalue(),
        {"inconcert_url": "inconcert/balanceador", "lead_url": "URL LEAD"},
        [result],
    )
    sheet = output.active
    headers = [cell.value for cell in sheet[1]]

    assert "URL LEAD" not in headers
    assert sheet["E2"].value == link
    assert sheet["E2"].hyperlink.target == link


def test_report_keeps_notice_when_program_was_not_preselected():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Programa", "URL"])
    sheet.append(["Maestría en Ingeniería de Datos e Infraestructura", "https://utel.test"])
    data = BytesIO()
    workbook.save(data)
    notice = "Incidencia corregida: UTEL abrió el campo Programa de interés sin preselección."
    result = {
        "row": {"sheet": "Sheet", "row_number": 2},
        "result": {
            "status": "PASS",
            "dry_run": False,
            "lead_url": "https://lead-balancer.scalahed.com/leads/detail/3141175",
            "lead_email": "qa@example.test",
            "summary": "Lead localizado en el Balanceador.",
            "program_selection_notice": notice,
            "stages": [],
        },
    }

    sheet = BotReportService().build(data.getvalue(), {}, [result]).active
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    assert notice in sheet.cell(2, headers["DETALLE BOT"]).value


def test_report_keeps_link_when_manage_email_validation_times_out():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nivel", "URL", "inconcert/balanceador"])
    sheet.append(["Licenciatura", "https://utel.test", None])
    data = BytesIO()
    workbook.save(data)
    link = "https://crm.test/mas/contact/people/view/123"
    result = {
        "row": {"sheet": "Sheet", "row_number": 2},
        "result": {
            "status": "FAIL",
            "dry_run": False,
            "lead_url": link,
            "lead_email": "qa@example.test",
            "summary": "Falló la validación visual",
            "stages": [{"status": "FAIL", "stage": "inconcert_manage", "message": "Timeout esperando el email"}],
        },
    }

    sheet = BotReportService().build(data.getvalue(), {"inconcert_url": "inconcert/balanceador"}, [result]).active

    assert sheet["C2"].value == link
    assert sheet["C2"].hyperlink.target == link
    assert sheet["D2"].value == "LEAD LOCALIZADO - VALIDACION PENDIENTE"
    assert "Faltó validar visualmente el email" in sheet["E2"].value
