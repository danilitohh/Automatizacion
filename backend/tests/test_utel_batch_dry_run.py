import asyncio
import io
import json
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from backend.app.api.routes import _run_utel_batch_job, cancel_utel_batch
from backend.app.automations.utel_inconcert.runner import UtelInconcertRunner, UtelQaError
from backend.app.config.settings import Settings
from backend.app.main import create_app


@pytest.fixture(autouse=True)
def crm_preflight(monkeypatch):
    """Impide acceso externo y permite comprobar cuándo se exige el preflight."""

    check = AsyncMock()
    monkeypatch.setattr(UtelInconcertRunner, "preflight_inconcert", check)
    return check


@pytest.mark.parametrize("dry_run", [True, False, None])
def test_batch_preserves_safe_mode_and_marks_excel(tmp_path, monkeypatch, crm_preflight, dry_run):
    seen = []

    async def fake_run(self, config):
        seen.append(config)
        return {"status": "PASS", "dry_run": config.dry_run, "stages": [], "summary": "Prueba local",
                "lead_url": None if config.dry_run else "https://crm.test/mas/contact/people/view/123",
                "utel_submission_attempted": not config.dry_run}

    monkeypatch.setattr(UtelInconcertRunner, "run", fake_run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    workbook.active.append(["Maestria", "https://example.test", "lateral", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    config = {"lead": {}}
    if dry_run is not None:
        config["dry_run"] = dry_run
    mapping = {"level": "Nivel", "utel_url": "URL", "form_type": "Location", "country": "Locale",
               "selected_sheet": "Sheet", "selected_row_number": 3}
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="test",
            inconcert_password="test",
            utel_test_phones_json='{"Mexico":["+525512345678"]}',
        )
    )
    with TestClient(app) as client:
        response = client.post('/api/bots/utel-inconcert/batch-run',
                               data={"config": json.dumps(config), "mapping": json.dumps(mapping)},
                               files={"file": ("test.xlsx", content.getvalue())})
        assert response.status_code == 202, response.text
        job_id = response.json()["job_id"]
        job = client.get(f'/api/bots/utel-inconcert/batch/{job_id}').json()
        assert job["status"] == "PASS", job
        assert job["completed"] == job["total"] == 1
        assert len(seen) == (2 if dry_run is False else 1)
        assert seen[0].dry_run is (dry_run is not False)
        if dry_run is False:
            assert seen[0].defer_crm_verification is True
            assert seen[1].verification_only is True
            crm_preflight.assert_awaited_once()
        else:
            crm_preflight.assert_not_awaited()
        assert seen[0].level == "Maestria"
        assert seen[0].source_filename == "test.xlsx"
        report = client.get(job["download_url"])
    sheet = load_workbook(io.BytesIO(report.content)).active
    assert sheet.cell(2, 5).value is None
    assert sheet.cell(3, 5).value == ("EXITOSO" if dry_run is False else "DRY RUN - NO ENVIADO")
    assert sheet.cell(1, 7).value == "URL LEAD"
    assert sheet.cell(3, 7).value == ("https://crm.test/mas/contact/people/view/123" if dry_run is False else None)
    if dry_run is False:
        assert sheet.cell(3, 7).hyperlink.target == sheet.cell(3, 7).value


def test_each_batch_row_uses_its_country_crm_not_a_stale_url(tmp_path, monkeypatch):
    seen = []

    async def fake_run(self, config):
        seen.append((config.country, config.inconcert_url))
        return {"status": "PASS", "dry_run": True, "stages": [], "summary": "Prueba local"}

    monkeypatch.setattr(UtelInconcertRunner, "run", fake_run)
    countries = [("México", "mas-utel"), ("Argentina", "mas-utel-arg"),
                 ("Colombia", "mas-utel-col"), ("Perú", "mas-utel-pe"),
                 ("Ecuador", "mas-utel-ec")]
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale", "CRM"])
    for country, _ in countries:
        workbook.active.append(["Licenciatura", "https://example.test", "footer", country, "https://wrong-country.test"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {"level": "Nivel", "utel_url": "URL", "form_type": "Location", "country": "Locale", "inconcert_url": "CRM"}
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="test",
            inconcert_password="test",
            utel_test_phones_json='{"Mexico":["+525512345678"]}',
        )
    )
    with TestClient(app) as client:
        response = client.post('/api/bots/utel-inconcert/batch-run',
                               data={"config": json.dumps({"lead": {}, "dry_run": True, "inconcert_url": "https://stale.test"}), "mapping": json.dumps(mapping)},
                               files={"file": ("test.xlsx", content.getvalue())})
        assert response.status_code == 202, response.text
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()
        assert job["status"] == "PASS", job
    assert seen == [(country, f"https://{tenant}.inconcertcc.com/login?redirect=%2Fmas%2Fhome") for country, tenant in countries]


def test_batch_can_run_only_the_rows_selected_for_retry(tmp_path, monkeypatch):
    seen = []

    async def fake_run(self, config):
        seen.append(config.level)
        return {"status": "PASS", "dry_run": True, "stages": [], "summary": "Prueba local"}

    monkeypatch.setattr(UtelInconcertRunner, "run", fake_run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    workbook.active.append(["Maestria", "https://example.test", "lateral", "Mexico"])
    workbook.active.append(["Doctorado", "https://example.test", "tarjeta", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {
        "level": "Nivel",
        "utel_url": "URL",
        "form_type": "Location",
        "country": "Locale",
        "selected_rows": [{"sheet": "Sheet", "row_number": 3}, {"sheet": "Sheet", "row_number": 4}],
    }
    app = create_app(Settings(database_path=tmp_path / "test.db", storage_dir=tmp_path / "storage", batch_delay_seconds=0))
    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={"config": json.dumps({"lead": {}, "dry_run": True}), "mapping": json.dumps(mapping)},
            files={"file": ("test.xlsx", content.getvalue())},
        )
        assert response.status_code == 202, response.text
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()
        assert job["status"] == "PASS", job
        assert job["total"] == job["completed"] == 2
    assert seen == ["Maestria", "Doctorado"]


def test_real_batch_stops_before_any_submission_without_crm_credentials(tmp_path, monkeypatch, crm_preflight):
    """No se crea ningún lead si después no sería posible reconciliarlo."""

    run = AsyncMock()
    monkeypatch.setattr(UtelInconcertRunner, "run", run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {
        "level": "Nivel",
        "utel_url": "URL",
        "form_type": "Location",
        "country": "Locale",
    }
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="",
            inconcert_password="",
            crm_username="",
            crm_password="",
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={"config": json.dumps({"lead": {}, "dry_run": False}), "mapping": json.dumps(mapping)},
            files={"file": ("test.xlsx", content.getvalue())},
        )
        assert response.status_code == 202
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()

    assert job["status"] == "FAIL"
    assert "faltan credenciales" in job["summary"].casefold()
    run.assert_not_awaited()
    crm_preflight.assert_not_awaited()


def test_batch_rejects_ambiguous_string_dry_run_before_starting(tmp_path, monkeypatch, crm_preflight):
    """La cadena ``false`` no puede saltarse el preflight y convertirse después."""

    run = AsyncMock()
    monkeypatch.setattr(UtelInconcertRunner, "run", run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    app = create_app(Settings(database_path=tmp_path / "test.db", storage_dir=tmp_path / "storage"))

    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={
                "config": json.dumps({"lead": {}, "dry_run": "false"}),
                "mapping": json.dumps({
                    "level": "Nivel",
                    "utel_url": "URL",
                    "form_type": "Location",
                    "country": "Locale",
                }),
            },
            files={"file": ("test.xlsx", content.getvalue())},
        )

    assert response.status_code == 400
    assert "dry_run debe ser booleano" in response.json()["detail"]
    run.assert_not_awaited()
    crm_preflight.assert_not_awaited()


def test_real_batch_stops_before_first_click_when_crm_preflight_fails(
    tmp_path, monkeypatch, crm_preflight
):
    """Una caída de login/Contactos aborta el lote completo antes de UTEL."""

    run = AsyncMock()
    monkeypatch.setattr(UtelInconcertRunner, "run", run)
    crm_preflight.side_effect = UtelQaError(
        "inconcert_preflight",
        "No se inició ningún envío: login de InConcert no disponible.",
    )
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {
        "level": "Nivel",
        "utel_url": "URL",
        "form_type": "Location",
        "country": "Locale",
    }
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="test",
            inconcert_password="test",
            utel_test_phones_json='{"Mexico":["+525512345678"]}',
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={"config": json.dumps({"lead": {}, "dry_run": False}), "mapping": json.dumps(mapping)},
            files={"file": ("test.xlsx", content.getvalue())},
        )
        assert response.status_code == 202
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()

    assert job["status"] == "FAIL"
    assert "no se inició ningún envío" in job["summary"].casefold()
    crm_preflight.assert_awaited_once()
    run.assert_not_awaited()


def test_real_batch_requires_enough_authorized_phones_before_first_click(
    tmp_path, monkeypatch, crm_preflight
):
    """El formato válido no sustituye al banco de números controlado por QA."""

    run = AsyncMock()
    monkeypatch.setattr(UtelInconcertRunner, "run", run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="test",
            inconcert_password="test",
            utel_test_phones_json="{}",
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={
                "config": json.dumps({"lead": {}, "dry_run": False}),
                "mapping": json.dumps({
                    "level": "Nivel",
                    "utel_url": "URL",
                    "form_type": "Location",
                    "country": "Locale",
                }),
            },
            files={"file": ("test.xlsx", content.getvalue())},
        )
        assert response.status_code == 202
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()

    assert job["status"] == "FAIL"
    assert "falta un banco de teléfonos autorizados" in job["summary"]
    crm_preflight.assert_not_awaited()
    run.assert_not_awaited()


def test_batch_merge_preserves_post_submit_warning_when_crm_finds_lead(tmp_path, monkeypatch):
    """La conciliacion exitosa conserva el aviso UTEL y nunca ejecuta una segunda fase de envio."""

    calls = []
    lead_url = "https://crm.test/mas/contact/people/view/456"
    original_warning = "Error al enviar. Contacta a soporte"

    async def fake_run(self, config):
        calls.append(config)
        if not config.verification_only:
            return {
                "status": "PASS",
                "dry_run": False,
                "summary": "Envio pendiente de conciliacion CRM.",
                "selected_program_name": "Licenciatura",
                "lead_email": config.lead.email,
                "lead_url": None,
                "utel_submission_attempted": True,
                "utel_submission": "pending",
                "utel_submission_message": original_warning,
                "lead_found": "pending",
                "stages": [],
                "screenshots": ["utel-warning.png"],
            }
        return {
            "status": "PASS",
            "dry_run": False,
            "summary": "Lead localizado en CRM.",
            "selected_program_name": "",
            "lead_email": config.lead.email,
            "lead_url": lead_url,
            "utel_submission_attempted": False,
            "utel_submission": "skipped",
            "utel_submission_message": "Envio ya realizado en la fase UTEL.",
            "inconcert_login": "success",
            "lead_found": "success",
            "lead_source": "inconcert",
            "conversion_found": "skipped",
            "stages": [],
            "screenshots": ["crm-lead.png"],
        }

    monkeypatch.setattr(UtelInconcertRunner, "run", fake_run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {
        "level": "Nivel",
        "utel_url": "URL",
        "form_type": "Location",
        "country": "Locale",
    }
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="test",
            inconcert_password="test",
            utel_test_phones_json='{"Mexico":["+525512345678"]}',
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={"config": json.dumps({"lead": {}, "dry_run": False}), "mapping": json.dumps(mapping)},
            files={"file": ("test.xlsx", content.getvalue())},
        )
        assert response.status_code == 202, response.text
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()
        report = client.get(job["download_url"])

    assert len(calls) == 2
    assert calls[0].defer_crm_verification is True
    assert calls[0].verification_only is False
    assert calls[1].verification_only is True
    final = job["results"][0]["result"]
    assert final["status"] == "PASS"
    assert final["lead_url"] == lead_url
    assert final["lead_found"] == "success"
    assert final["utel_submission"] == "success"
    assert original_warning in final["utel_submission_message"]
    assert final["selected_program_name"] == "Licenciatura"
    assert final["screenshots"] == ["utel-warning.png", "crm-lead.png"]

    sheet = load_workbook(io.BytesIO(report.content)).active
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    assert sheet.cell(2, headers["RESULTADO FORMULARIO"]).value == "EXITOSO"
    assert original_warning in sheet.cell(2, headers["DETALLE BOT"]).value
    link_cell = sheet.cell(2, headers["URL LEAD"])
    assert link_cell.value == lead_url
    assert link_cell.hyperlink.target == lead_url


def test_batch_is_failed_when_crm_does_not_find_clicked_submission(tmp_path, monkeypatch):
    """Un lote no puede declararse exitoso si CRM no confirma el unico envio."""

    calls = []
    original_notice = "Formulario enviado y confirmado correctamente."

    async def fake_run(self, config):
        calls.append(config)
        if not config.verification_only:
            return {
                "status": "PASS",
                "dry_run": False,
                "summary": "Formulario enviado; pendiente de CRM.",
                "lead_url": None,
                "utel_submission_attempted": True,
                "utel_submission": "success",
                "utel_submission_message": original_notice,
                "lead_found": "pending",
                "stages": [],
                "screenshots": ["utel-success.png"],
            }
        return {
            "status": "FAIL",
            "dry_run": False,
            "summary": "Lead no encontrado en CRM.",
            "lead_url": None,
            "utel_submission_attempted": False,
            "utel_submission": "skipped",
            "utel_submission_message": "Envio ya realizado en la fase UTEL.",
            "inconcert_login": "success",
            "lead_found": "failed",
            "conversion_found": "skipped",
            "stages": [],
            "screenshots": ["crm-missing.png"],
        }

    monkeypatch.setattr(UtelInconcertRunner, "run", fake_run)
    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://example.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {
        "level": "Nivel",
        "utel_url": "URL",
        "form_type": "Location",
        "country": "Locale",
    }
    app = create_app(
        Settings(
            database_path=tmp_path / "test.db",
            storage_dir=tmp_path / "storage",
            batch_delay_seconds=0,
            inconcert_username="test",
            inconcert_password="test",
            utel_test_phones_json='{"Mexico":["+525512345678"]}',
        )
    )

    with TestClient(app) as client:
        response = client.post(
            "/api/bots/utel-inconcert/batch-run",
            data={"config": json.dumps({"lead": {}, "dry_run": False}), "mapping": json.dumps(mapping)},
            files={"file": ("test.xlsx", content.getvalue())},
        )
        assert response.status_code == 202, response.text
        job = client.get(f'/api/bots/utel-inconcert/batch/{response.json()["job_id"]}').json()
        report = client.get(job["download_url"])

    assert len(calls) == 2
    assert calls[1].verification_only is True
    final = job["results"][0]["result"]
    assert job["status"] == "FAIL"
    assert job["success"] == 0
    assert job["failed"] == 1
    assert final["status"] == "FAIL"
    assert final["lead_found"] == "failed"
    assert final["lead_url"] is None
    assert original_notice in final["utel_submission_message"]
    assert "no se reenv" in final["utel_submission_message"].casefold()
    assert final["screenshots"] == ["utel-success.png", "crm-missing.png"]

    sheet = load_workbook(io.BytesIO(report.content)).active
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    assert sheet.cell(2, headers["RESULTADO FORMULARIO"]).value == "ERROR"
    assert original_notice in sheet.cell(2, headers["DETALLE BOT"]).value
    assert sheet.cell(2, headers["URL LEAD"]).value is None


def test_batch_cancel_finishes_current_row_and_reconciles_before_stopping(
    tmp_path, monkeypatch
):
    """Detener nunca pierde un clic ni inicia la fila siguiente."""

    workbook = Workbook()
    workbook.active.append(["Nivel", "URL", "Location", "Locale"])
    workbook.active.append(["Licenciatura", "https://first.test", "footer", "Mexico"])
    workbook.active.append(["Maestria", "https://second.test", "footer", "Mexico"])
    content = io.BytesIO()
    workbook.save(content)
    mapping = {
        "level": "Nivel",
        "utel_url": "URL",
        "form_type": "Location",
        "country": "Locale",
    }
    settings = Settings(
        database_path=tmp_path / "test.db",
        storage_dir=tmp_path / "storage",
        batch_delay_seconds=0,
        inconcert_username="test",
        inconcert_password="test",
        utel_test_phones_json='{"Mexico":["+525512345678","+525598765432"]}',
    )
    job_id = "cancel-safe"
    state = SimpleNamespace(
        settings=settings,
        utel_batch_jobs={
            job_id: {
                "job_id": job_id,
                "status": "RUNNING",
                "total": 2,
                "completed": 0,
                "success": 0,
                "failed": 0,
                "pending": 0,
                "cancel_requested": False,
            }
        },
        bot_tasks={},
    )
    application = SimpleNamespace(state=state)
    request = SimpleNamespace(app=application)

    async def scenario():
        started = asyncio.Event()
        release = asyncio.Event()
        calls = []

        async def fake_run(self, config):
            calls.append(config)
            if config.verification_only:
                return {
                    "status": "PASS",
                    "dry_run": False,
                    "summary": "Lead localizado.",
                    "lead_url": "https://crm.test/mas/contact/people/view/789",
                    "lead_found": "success",
                    "utel_submission_attempted": False,
                    "stages": [],
                    "screenshots": [],
                }
            started.set()
            await release.wait()
            return {
                # Demuestra que se encola por clic, no por el estado visual.
                "status": "FAIL",
                "dry_run": False,
                "summary": "Respuesta visual no interpretable.",
                "lead_url": None,
                "lead_found": "pending",
                "utel_submission": "pending",
                "utel_submission_message": "Aviso posterior al clic.",
                "utel_submission_attempted": True,
                "stages": [],
                "screenshots": [],
            }

        monkeypatch.setattr(UtelInconcertRunner, "run", fake_run)
        task = asyncio.create_task(
            _run_utel_batch_job(
                application,
                job_id,
                content.getvalue(),
                "cancel.xlsx",
                {"lead": {}, "dry_run": False},
                mapping,
            )
        )
        state.bot_tasks[job_id] = task
        await started.wait()
        cancellation = await cancel_utel_batch(request, job_id)
        status_when_requested = cancellation["status"]
        release.set()
        await task
        return calls, status_when_requested

    calls, status_when_requested = asyncio.run(scenario())

    job = state.utel_batch_jobs[job_id]
    assert status_when_requested == "RUNNING"
    assert job["status"] == "CANCELLED"
    assert job["completed"] == 1
    assert len(calls) == 2
    assert calls[0].utel_url == "https://first.test"
    assert calls[1].verification_only is True
    assert job["results"][0]["result"]["lead_url"].endswith("/789")
    assert "conciliados" in job["summary"]
